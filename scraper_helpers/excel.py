import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from scraper_models.constants import HEADERS

# ============================================================
# FOR FUTURE MODIFICATIONS OF HEADERS OR COLUMNS
# ============================================================
rank_row_value = 5
app_name_row_value = 7
icon_preview_row_value = 8
app_link_row_value = 9
source_row_value = 10
icon_path_row_value = 11

# ============================================================
# TEXT-ONLY APPENDER (NO ICONS HERE)
# ============================================================

def append_rows(ws, rows):
    """
    Append rows as plain text only.
    Icon handling is done in append_rows_to_category_sheets.
    """
    for r_idx, row in enumerate(rows, start=ws.max_row + 1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

# ============================================================
# WORKBOOK PREPARATION
# ============================================================

def prepare_workbook_for_append(
    output_file: str,
    headers: list,
    category_sheets=("Music", "Navigation", "Messaging")
):
    out_dir = os.path.dirname(output_file)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    if not os.path.exists(output_file):
        Workbook().save(output_file)

    wb = load_workbook(output_file)
    bold_font = Font(bold=True)

    ws_map = {}
    for name in category_sheets:
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb.create_sheet(name)

        # Ensure headers
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = bold_font

        ws_map[name] = ws

    # Remove default Sheet
    if "Sheet" in wb.sheetnames and "Sheet" not in ws_map:
        ws0 = wb["Sheet"]
        if ws0.max_row <= 1 and ws0.max_column <= 1:
            wb.remove(ws0)

    # Icon column width 
    icon_col_letter = get_column_letter(len(headers))
    for ws in ws_map.values():
        ws.column_dimensions[icon_col_letter].width = 11

    # Preview column width 
    preview_col_letter = get_column_letter(len(headers) + 1)
    for ws in ws_map.values():
        ws.column_dimensions[preview_col_letter].width = 14

    return wb, ws_map


# ============================================================
# CATEGORY APPENDER WITH URL-FIRST ICON LOGIC
# ============================================================

def append_rows_to_category_sheets(
    ws_map: dict,
    rows_out: list,
    file_category: str,
    *,
    input_dir=None,
    base_url: str = "",
    icon_lookup=None,
    icon_size_px=30,
):
    yellow_fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid"
    )
    
    icon_col_letter = get_column_letter(icon_path_row_value)
    preview_col_letter = get_column_letter(icon_preview_row_value)

    def _is_http_url(s: str) -> bool:
        return bool(s and s.lower().startswith(("http://", "https://")))

    def _to_hyperlink(source_path: str):
        if not source_path:
            return None, ""
        sp_norm = os.path.abspath(source_path)
        if base_url:
            try:
                rel = os.path.relpath(
                    sp_norm,
                    start=os.path.abspath(input_dir or os.getcwd())
                )
                rel_url = "/".join(rel.split(os.sep))
                return f"{base_url.rstrip('/')}/{rel_url}", sp_norm
            except Exception:
                pass
        return "file:///" + sp_norm.replace("\\", "/"), sp_norm

    for row in rows_out:
        if file_category not in ws_map:
            continue

        ws = ws_map[file_category]
        i = ws.max_row + 1

        # ------------------------------
        # Write row values
        # ------------------------------
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)

            # App Link 
            if j == app_link_row_value and isinstance(val, str) and _is_http_url(val):
                cell.hyperlink = val
                cell.style = "Hyperlink"

            # Source 
            if j == source_row_value and isinstance(val, str) and val.strip():
                url, display = _to_hyperlink(val)
                if url:
                    cell.value = display
                    cell.hyperlink = url
                    cell.style = "Hyperlink"

        # ------------------------------
        # ICON: URL FIRST, EMBED FALLBACK
        # ------------------------------
        if callable(icon_lookup):
            app_name = row[app_name_row_value - 1] if len(row) >= (app_name_row_value - 1) else ""
            app_link = row[app_link_row_value - 1] if len(row) >= (app_link_row_value - 1) else ""

            try:
                kind, value = icon_lookup(app_link, app_name)
            except Exception:
                kind, value = None, None

            # URL → hyperlink
            if kind == "url" and value:
                cell = ws.cell(row=i, column=icon_path_row_value, value=value)
                cell.hyperlink = value
                cell.style = "Hyperlink"

            # EMBED → image
            elif kind == "embed" and value:
                try:
                    img = XLImage(value)
                    img.width = icon_size_px
                    img.height = icon_size_px
                    ws.add_image(img, f"{icon_col_letter}{i}")
                    ws.row_dimensions[i].height = max(
                        ws.row_dimensions[i].height or 0,
                        icon_size_px / 0.75
                    )
                except Exception:
                    pass

        # ------------------------------
        # ICON PREVIEW FORMULA (COLUMN K)
        # ------------------------------
        try:
            ws.cell(
                row=i,
                column=icon_preview_row_value,
                value=f"=IMAGE({icon_col_letter}{i})"
            )
        except Exception:
            pass

        # ------------------------------
        # Highlight Rank == 1
        # ------------------------------
        try:
            rank_int = int(str(row[rank_row_value - 1]).strip())
        except Exception:
            rank_int = None

        if rank_int == 1:
            for col_idx in range(1, icon_preview_row_value + 1):
                ws.cell(row=i, column=col_idx).fill = yellow_fill